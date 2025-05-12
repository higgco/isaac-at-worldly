import psycopg2
import time
import os
from dotenv import load_dotenv
load_dotenv()
import ast
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.drawing.image import Image
import pandas as pd
from datetime import datetime
from matplotlib.ticker import MaxNLocator
print("Starting script...")  # Debugging output
# Database connection parameters
def get_connection():
    return psycopg2.connect(
        host=os.getenv('DB_HOST'),
        dbname=os.getenv('DB_NAME'),
        user=os.getenv('DB_USER'),
        password=os.getenv('DB_PASSWORD'),
        port=int(os.getenv('DB_PORT', 5432)),
        connect_timeout=10
    )
print("Connection successful!")  # Debugging output
retry_count = 0
max_retries = 5
delay_interval = 10  # in seconds
brand = "ALL FEM23_24"
query = f'''
SELECT * FROM isaac_hopwood.fem2023_24_answers_final
'''
# Unit conversion function with error handling
def convert_to_kg(row):
    quantity = row['quantity']
    # Check if quantity is None
    if quantity is None:
        return None, row['unit']  # Return None and the original unit if quantity is not valid
    try:
        quantity = float(quantity)
    except ValueError:
        # Handle case where quantity can't be converted to float
        print(f"Error converting quantity: {row['quantity']}")
        return None, row['unit']  # Return None and the original unit if conversion fails
    unit = row['unit']
    conversion_factors = {
        'g': 0.001,          # 1 gram = 0.001 kg
        'kg': 1,             # 1 kg = 1 kg
        'lb': 0.453592,      # 1 pound = 0.453592 kg
        'oz': 0.0283495,     # 1 ounce = 0.0283495 kg
        'tonmetric': 1000,   # 1 metric ton = 1000 kg
        'tonshort': 907.18474 # 1 short ton = 907.18474 kg
    }
    # Convert the quantity to kg
    converted_quantity = quantity * conversion_factors.get(unit, 1)
    # Update the unit to reflect the conversion, but keep it as 'kg' if it's already in kg
    if unit == 'kg':
        new_unit = 'kg'
    else:
        new_unit = f"kg (converted from {unit})"
    # Return the converted quantity and updated unit
    return converted_quantity, new_unit
# Function to process the table and combine quantities into a single column
def process_table(input_table, column_names):
    processed_list = []
    waste_sources = [
        "textile", "leather", "rubber", "metal", "plastic", "paper", "cans", "wood", "food", 
        "glass", "cartons", "foams", "wastewatertreatmentsludge", "nh_general", "slag_nh",
        "nh_other", "prodchemdrum", "prodfilmprint", "prodsludge", "prodchem", "prodcompgas", 
        "prodcontammat", "dombatteries", "domflolight", "dominkcart", "domoilgrease", 
        "productionoil", "metalsludge", "h_emptyother", "domelectronic", "domcoalcomb", 
        "slag_h", "h_other"
    ]
    for row in input_table:
        row_dict = dict(zip(column_names, row))
        assessment_id = row_dict.get('assessment_id')
        for source in waste_sources:
            method_col = f"{source}_method"
            quantity_col = f"{source}_quant"
            unit_col = f"{source}_unit"
            method = row_dict.get(method_col)
            quantity = row_dict.get(quantity_col)
            unit = row_dict.get(unit_col)
            if quantity is not None and quantity != 0:
                try:
                    methods = ast.literal_eval(method) if method else []
                except (ValueError, SyntaxError):
                    methods = [method] if method else []
                if not methods:
                    methods = ['Unknown']
                split_quantity = quantity / len(methods) if methods else 0
                add_estimate = len(methods) > 1
                for single_method in methods:
                    processed_row = {
                        'assessment_id': assessment_id,
                        'waste_source': source,
                        'disposal_method': single_method,
                        'quantity': split_quantity,
                        'unit': unit,
                        'estimated': add_estimate
                    }
                    processed_list.append(processed_row)
    return processed_list
# Get the script's filename and create a unique output filename
script_filename = os.path.basename(__file__)
base_filename = os.path.splitext(script_filename)[0]
# Set the file paths
output_directory = f"C:\\YOUR PATH HERE\\{brand} Results"
excel_output_file = os.path.join(output_directory, f"{base_filename} results.xlsx")
png_image_file = os.path.join(output_directory, f'{excel_output_file} bar chart and table.png')
# Create the output directory if it doesn't exist
os.makedirs(output_directory, exist_ok=True)
# Attempt to execute the query with retries
while retry_count < max_retries:
    connection = None
    try:
        connection = get_connection()
        connection.set_isolation_level(psycopg2.extensions.ISOLATION_LEVEL_READ_COMMITTED)
        with connection.cursor() as cursor:
            print(f"Attempt {retry_count + 1}: Executing query...")
            cursor.execute("SET statement_timeout = 300000;")
            cursor.execute(query)
            results = cursor.fetchall()
            column_names = [desc[0] for desc in cursor.description]
            if results:
                processed_results = []
                for row in results:
                    new_row = list(row)
                    for i in range(len(column_names)):
                        if column_names[i].endswith('quant'):
                            unit_index = column_names.index(column_names[i].replace('quant', 'unit'))
                            converted_quantity, new_unit = convert_to_kg({'quantity': new_row[i], 'unit': new_row[unit_index]})
                            new_row[i] = converted_quantity
                            new_row[unit_index] = new_unit
                    processed_results.append(new_row)
                processed_list = process_table(processed_results, column_names)
                # Create DataFrame from processed_list
                processed_list_df = pd.DataFrame(processed_list)
                # Print debugging information
                print("processed_list_df shape:", processed_list_df.shape)
                print("processed_list_df columns:", processed_list_df.columns)
                print("processed_list_df head:", processed_list_df.head())
                # Create pivot table
                pivot_df = processed_list_df.pivot_table(
                    values='quantity',
                    index='disposal_method',
                    columns='estimated',
                    aggfunc='sum',
                    fill_value=0
                )
                # Print debugging information for pivot_df
                print("Initial pivot_df shape:", pivot_df.shape)
                print("Initial pivot_df columns:", pivot_df.columns)
                # Check the number of columns before renaming
                if pivot_df.shape[1] == 2:
                    # Rename columns in the pivot table
                    pivot_df.columns = ['Not estimated', 'estimated']
                else:
                    print(f"Expected 2 columns in pivot_df, but found {pivot_df.shape[1]} columns.")
                    continue  # Skip this iteration if the structure is incorrect
                # Print debugging information after renaming
                print("pivot_df shape after renaming:", pivot_df.shape)
                print("pivot_df columns after renaming:", pivot_df.columns)
                try:
                    # Save the data and pivot table to Excel
                    with pd.ExcelWriter(excel_output_file, engine='xlsxwriter') as writer:
                        # Save the processed_list_df to the 'Data' sheet
                        processed_list_df = pd.DataFrame(processed_list)
                        processed_list_df.to_excel(writer, sheet_name='Data', index=False)
                        # Prepare and save the pivot table to the 'PivotTable' sheet
                        pivot_df = processed_list_df.pivot_table(
                            values='quantity',
                            index='disposal_method',
                            columns='estimated',
                            aggfunc='sum',
                            fill_value=0
                        )
                        # Rename columns in the pivot table
                        pivot_df.columns = ['Not estimated', 'estimated']
                        if pivot_df.shape[1] == 2:
                            pivot_df.columns = ['Not estimated', 'estimated']
                        else:
                            print(f"Expected 2 columns in pivot_df, but found {pivot_df.shape[1]} columns.")
                            continue  # Skip this iteration if the structure is incorrect
                        print("pivot_df columns after renaming:", pivot_df.columns)
                        # Save the pivot table to the 'PivotTable' sheet
                        pivot_df.to_excel(writer, sheet_name='PivotTable')
                    print(f"Data and pivot table saved to {excel_output_file}")
                    break
                except PermissionError:
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    backup_file = os.path.join(output_directory, f"{base_filename} results ({timestamp}).xlsx")
                    print(f"Permission denied for {excel_output_file}. Saving results to backup file {backup_file}.")
                    with pd.ExcelWriter(backup_file, engine='xlsxwriter') as writer:
                        processed_list_df.to_excel(writer, sheet_name='Data', index=False)
                        pivot_df.to_excel(writer, sheet_name='PivotTable')
                    print(f"Results saved to backup file {backup_file}")
                    excel_output_file = backup_file
                    break
            else:
                print("Query executed successfully, but no data was returned.")
            connection.commit()
            break
    except psycopg2.OperationalError as e:
        print(f"Operational error encountered: {e}")
        retry_count += 1
        if retry_count >= max_retries:
            print(f"Query failed after {max_retries} retries.")
            break
        print(f"Retrying in {delay_interval} seconds...")
        time.sleep(delay_interval)
    except psycopg2.Error as e:
        if e.pgcode == '40001':
            print(f"Error 40001 encountered. Retrying... ({retry_count + 1})")
            retry_count += 1
            if connection is not None:
                try:
                    connection.rollback()
                except psycopg2.OperationalError as rollback_error:
                    print(f"Rollback error: {rollback_error}")
            time.sleep(delay_interval)
        else:
            print(f"An unexpected error occurred: {e}")
            if connection is not None:
                try:
                    connection.rollback()
                except psycopg2.OperationalError as rollback_error:
                    print(f"Rollback error: {rollback_error}")
            break
    finally:
        if connection is not None:
            try:
                connection.close()
            except Exception as close_error:
                print(f"Error closing connection: {close_error}")
# Plotting
# Convert quantities from kg to metric tons (MT)
pivot_df_mt = pivot_df / 1e3  # Divide by 1000 to convert from kg to metric tons (MT)
# Set the figure size to match 2256x1092 pixels at 100 DPI
fig = plt.figure(figsize=(22.56, 10.92), dpi=100)
# Create the stacked bar plot with metric tons
ax = fig.add_subplot(121)  # Left subplot for the bar plot
bars1 = ax.bar(pivot_df_mt.index, pivot_df_mt['Not estimated'], label='Not estimated', color='#1f77b4')
bars2 = ax.bar(pivot_df_mt.index, pivot_df_mt['estimated'], bottom=pivot_df_mt['Not estimated'], label='estimated', color='#ff7f0e')
# Set labels and title
ax.set_ylabel('Amount (Metric Tons)', fontsize=12)
ax.set_title(f'{brand} FEM23 Facility Waste by Disposal Method', fontsize=14, fontweight='bold')
# Automatically adjust y-axis ticks to avoid excessive ticks
ax.yaxis.set_major_locator(MaxNLocator(integer=True))  # Automatic tick calculation for large ranges
# Format y-axis labels to show values in metric tons
def metric_tons_formatter(x, pos):
    return f'{x:.0f} MT'
ax.yaxis.set_major_formatter(plt.FuncFormatter(metric_tons_formatter))
# Add horizontal gridlines
ax.yaxis.grid(True, linestyle='--', alpha=0.7)
# Customize x-axis
plt.xticks(rotation=45, ha='right', fontsize=10)
ax.set_xlim(-0.5, len(pivot_df_mt.index) - 0.5)  # Adjust x-axis limits
# Add a horizontal line at y=0
ax.axhline(y=0, color='k', linewidth=0.8)
# Customize legend
ax.legend(title='Data Type', title_fontsize='12', fontsize='10', loc='upper left')
# Create table data (totals and values for each bar) in metric tons
table_data = []
for index in pivot_df_mt.index:
    not_estimated = pivot_df_mt['Not estimated'][index]
    estimated = pivot_df_mt['estimated'][index]
    total = not_estimated + estimated
    table_data.append([f'{not_estimated:.0f} MT', f'{estimated:.0f} MT', f'{total:.0f} MT'])
# Create table columns and rows
columns = ['Not estimated', 'estimated', 'Total']
rows = pivot_df_mt.index
# Add table to the right subplot
table_ax = fig.add_subplot(122)  # Right subplot for the table
table_ax.axis('tight')  # Ensure the table is tight
table_ax.axis('off')  # Hide axes for the table
# Create the table
the_table = table_ax.table(cellText=table_data,
                            colLabels=columns,
                            rowLabels=rows,
                            loc='center')
# Adjust table properties for better appearance
the_table.auto_set_font_size(True)
the_table.scale(1, 1)  # Adjust the size as needed
# Set positions to avoid overlap
ax.set_position([0.1, 0.1, 0.65, 0.8])  # Adjust left, bottom, width, height for the bar plot
table_ax.set_position([0.8, 0.1, 0.15, 0.8])  # Adjust left, bottom, width, height for the table
# Adjust layout to make room for both plots
plt.subplots_adjust(left=0.05, right=0.95)  # Overall figure adjustments
# Step 1: Save the plot as PNG
fig.savefig(png_image_file)
# Step 2: Open the existing Excel file using openpyxl
wb = openpyxl.load_workbook(excel_output_file)
# Step 3: Create a new sheet for the image if it doesn't exist
if 'Chart' not in wb.sheetnames:
    wb.create_sheet('Chart')
# Load the newly created sheet
sheet = wb['Chart']
# Step 4: Insert the PNG image into the new sheet
img = Image(png_image_file)
# Position the image
img.anchor = 'A1'
# Add the image to the sheet
sheet.add_image(img)
# Step 5: Save the changes to the workbook
wb.save(excel_output_file)
# Cleanup (optional): remove the temporary PNG file after saving
os.remove(png_image_file)
os.startfile(excel_output_file)
print('Interactive plot open...')
plt.show()
print("Fin.")